from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
import re

from app import (
    arquivo_consemavi_padrao,
    arquivo_modelo_padrao,
    arquivo_word_padrao,
    exit_dir,
    temp_dir,
    meses,
    gerar_relatorio_final,
)

import base64

def carregar_imagem_base64(caminho):
    with open(caminho, "rb") as f:
        return base64.b64encode(f.read()).decode()

img_base64 = carregar_imagem_base64("img\logo.png")

st.set_page_config(page_title="Monitor de Metas Institucionais", layout="wide")

st.markdown("""
<style>
.stApp { background-color: #d1d5db; }
section[data-testid="stSidebar"] {
    background-color: #fdf2f8 !important;
    border-right: 2px solid #fce7f3;
}
.stSidebar label {
    color: #831843 !important;
    font-weight: bold !important;
}
.stButton > button {
    background-color: #ff007f !important;
    color: white !important;
    border-radius: 12px !important;
    width: 100% !important;
    font-weight: 800 !important;
    padding: 15px !important;
    border: none !important;
}
.stDownloadButton > button {
    background-color: #831843 !important;
    color: white !important;
    border-radius: 12px !important;
    width: 100% !important;
    font-weight: 800 !important;
    padding: 15px !important;
    border: none !important;
}
div[data-testid="stFileUploaderDropzoneInstructions"] { display: none !important; }
div[data-testid="stFileUploaderDropzone"] small { display: none !important; }
div[data-testid="stFileUploaderDropzone"] p { display: none !important; }

.header-banner {
    background: linear-gradient(90deg, #ff007f 0%, #ff69b4 100%);
    padding: 20px;
    border-radius: 15px;
    text-align: center;
    color: white;
    margin-bottom: 25px;
}
.preview-wrapper {
    background: white;
    border-radius: 12px;
    padding: 14px;
    box-shadow: 0 8px 18px rgba(0,0,0,0.08);
    overflow-x: auto;
    overflow-y: hidden;
    width: 100%;
}
.preview-table {
    border-collapse: collapse;
    font-size: 12px;
    background: white;
    table-layout: fixed;
    min-width: 1100px;
}
.preview-table th,
.preview-table td {
    border: 1px solid #e5e7eb;
    padding: 4px 6px;
    text-align: left;
    max-width: 140px;
    min-width: 90px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}
.preview-table th {
    background-color: #fdf2f8;
    color: #831843;
    font-weight: 700;
    position: sticky;
    top: 0;
    z-index: 1;
}
.preview-wrapper::-webkit-scrollbar {
    height: 12px;
}
.preview-wrapper::-webkit-scrollbar-track {
    background: #f3f4f6;
    border-radius: 999px;
}
.preview-wrapper::-webkit-scrollbar-thumb {
    background: #ff69b4;
    border-radius: 999px;
}
.preview-wrapper::-webkit-scrollbar-thumb:hover {
    background: #ff007f;
}
</style>
""", unsafe_allow_html=True)


def salvar_upload(uploaded_file, pasta_destino: Path) -> Path:
    pasta_destino.mkdir(exist_ok=True, parents=True)
    destino = pasta_destino / uploaded_file.name
    with open(destino, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return destino


def carregar_preview_excel_completo(caminho_excel: Path):
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    nome_aba = next((nome for nome in wb.sheetnames if "49" in str(nome)), wb.sheetnames[0])
    ws = wb[nome_aba]

    matriz = []
    for r in range(1, ws.max_row + 1):
        linha = []
        for c in range(1, ws.max_column + 1):
            valor = ws.cell(r, c).value
            linha.append("" if valor is None else valor)
        matriz.append(linha)

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = merged_range.bounds
        valor_principal = ws.cell(min_row, min_col).value
        valor_principal = "" if valor_principal is None else valor_principal

        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                matriz[r - 1][c - 1] = valor_principal

    df = pd.DataFrame(matriz)
    df = df.loc[:, ~(df.apply(lambda col: col.astype(str).str.strip().eq("").all(), axis=0))]
    df = df.loc[~(df.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1))]
    df = df.reset_index(drop=True)

    return df, nome_aba

def formatar_numero_br(valor):
    if pd.isna(valor) or valor == "":
        return ""

    # número real
    if isinstance(valor, (int, float)):
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    texto = str(valor).strip()

    # se já está formatado em brasileiro, mantém
    if re.fullmatch(r"\d{1,3}(\.\d{3})*,\d{2}", texto):
        return texto

    # tenta converter texto para número
    try:
        if "," in texto:
            num = float(texto.replace(".", "").replace(",", "."))
        else:
            num = float(texto)
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return texto
    
def dataframe_para_html_com_cabecalho(df: pd.DataFrame):
    # Remove linhas que sejam apenas "Total" repetido (limpeza visual)
    df = df.loc[
        ~df.apply(
            lambda row: row.astype(str).str.lower().str.contains("total").sum() > 6,
            axis=1
        )
    ]

    if df.empty:
        return '<div class="preview-wrapper">Tabela vazia</div>'

    # Define cabeçalho e corpo baseado na estrutura da planilha Meta 49
    if len(df) >= 3:
        cabecalho = df.iloc[2].fillna("").tolist()
        corpo = df.iloc[3:].copy()
    else:
        cabecalho = df.iloc[0].fillna("").tolist()
        corpo = df.iloc[1:].copy()

    html = ['<div class="preview-wrapper"><table class="preview-table">']

    # Gerar Cabeçalho
    html.append("<thead><tr>")
    for valor in cabecalho:
        texto = "" if pd.isna(valor) else str(valor)
        html.append(f"<th>{texto}</th>")
    html.append("</tr></thead>")

    # Gerar Corpo com formatação numérica brasileira
    html.append("<tbody>")
    for _, row in corpo.iterrows():
        html.append("<tr>")
        for valor in row:
            # AQUI ESTÁ A MUDANÇA: Usamos a sua função formatar_numero_br
            texto = formatar_numero_br(valor)
            html.append(f"<td>{texto}</td>")
        html.append("</tr>")
    html.append("</tbody>")
    
    html.append("</table></div>")
    return "".join(html)

with st.sidebar:
    st.markdown("<h2 style='color:#ff007f; text-align:center;'> Painel</h2>", unsafe_allow_html=True)

    ano_sel = st.selectbox(" Selecione o ano", options=[2026, 2025], index=0)

    opcoes_meses = ["Todos os meses"] + meses
    mes_sel = st.selectbox(" Selecione o mês", options=opcoes_meses, index=1)

    st.divider()

    st.caption("Se não enviar arquivo, o app usa os arquivos padrão da pasta data.")
    word_file = st.file_uploader(" Carregar word convias (docx)", type=["docx"])
    excel_file = st.file_uploader(" Carregar excel consemavi (xlsx)", type=["xlsx"])

    st.divider()
    btn_gerar = st.button(" Gerar relatório")

st.markdown(f"""
<div class="header-banner" style="
    display:flex;
    align-items:center;
    justify-content:center;
    gap:15px;
    padding:20px;
    background: linear-gradient(90deg, #ff007f, #ff69b4);
    border-radius:12px;
">
    <img src="data:image/png;base64,{img_base64}" width="120">
    <div>
        <h1 style="margin:0; color:white; align-items:center;"> Monitor de Metas Municipais</h1>
        <p style="margin:0; opacity:0.7; color:white;">Coplan-Dados</p>
    </div>
</div>
""", unsafe_allow_html=True)



if btn_gerar:
    with st.spinner("Lendo arquivos, validando somas e preenchendo o modelo..."):
        try:
            caminho_word = salvar_upload(word_file, temp_dir) if word_file else arquivo_word_padrao
            caminho_excel = salvar_upload(excel_file, temp_dir) if excel_file else arquivo_consemavi_padrao
            caminho_saida = exit_dir / "meta49_preenchido.xlsx"

            caminho_final = gerar_relatorio_final(
                ano_ref=ano_sel,
                mes_ref=mes_sel,
                caminho_word=caminho_word,
                caminho_consemavi=caminho_excel,
                caminho_saida=caminho_saida,
            )

            caminho_final = Path(caminho_final)

            if caminho_final.exists():
                st.success("Excel atualizado com sucesso!")
                st.info(f"Arquivo salvo em: {caminho_final.absolute()}")
                st.session_state["Arquivo_gerado"] = str(caminho_final)
            else:
                st.error("Erro: arquivo gerado não encontrado.")

        except Exception as e:
            st.error(f"Erro ao gerar relatório: {e}")

if "arquivo_gerado" in st.session_state:
    caminho_preview = Path(st.session_state["arquivo_gerado"])

    if caminho_preview.exists():
        try:
            df_preview, aba_nome = carregar_preview_excel_completo(caminho_preview)

            st.markdown(f"### Preview do arquivo gerado - aba `{aba_nome}`")

            html_tabela = dataframe_para_html_com_cabecalho(df_preview)
            st.markdown(html_tabela, unsafe_allow_html=True)

            st.divider()

            with open(caminho_preview, "rb") as f:
                st.download_button(
                    label="Baixar excel gerado",
                    data=f,
                    file_name=caminho_preview.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"Erro ao exibir o arquivo gerado: {e}")
else:
    st.info("Selecione o período e clique em gerar relatório.")